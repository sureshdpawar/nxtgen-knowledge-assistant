from pathlib import Path

from openpyxl import (
    load_workbook,
)

from app.parsers.base import (
    BaseParser,
    ParsedResult,
)


class XlsxParser(
    BaseParser
):

    MAX_EMPTY_ROWS = 25

    def extract(
        self,
        file_path: Path,
    ) -> ParsedResult:

        workbook = (
            load_workbook(
                filename=str(
                    file_path
                ),
                read_only=True,
                data_only=True,
            )
        )

        pages: list[dict] = []

        try:
            for (
                sheet_index,
                worksheet,
            ) in enumerate(
                workbook.worksheets,
                start=1,
            ):

                text_parts: list[str] = [
                    (
                        "Worksheet: "
                        f"{worksheet.title}"
                    )
                ]

                consecutive_empty_rows = 0

                for row in (
                    worksheet
                    .iter_rows(
                        values_only=True
                    )
                ):

                    values = [
                        self._format_value(
                            value
                        )
                        for value
                        in row
                    ]

                    while (
                        values
                        and values[-1] == ""
                    ):
                        values.pop()

                    if not any(
                        value
                        for value
                        in values
                    ):
                        consecutive_empty_rows += 1

                        if (
                            consecutive_empty_rows
                            >= self.MAX_EMPTY_ROWS
                        ):
                            #
                            # Avoid walking through
                            # thousands of styled but
                            # empty rows.
                            #
                            continue

                        continue

                    consecutive_empty_rows = 0

                    text_parts.append(
                        " | ".join(
                            values
                        )
                    )

                pages.append(
                    {
                        "page":
                            sheet_index,

                        "text":
                            "\n".join(
                                text_parts
                            ),
                    }
                )

            return {
                "pages":
                    pages,

                "metadata": {
                    "page_count":
                        len(
                            workbook
                            .worksheets
                        ),

                    "sheet_count":
                        len(
                            workbook
                            .worksheets
                        ),

                    "sheet_names": [
                        worksheet.title
                        for worksheet
                        in workbook
                        .worksheets
                    ],

                    "document_type":
                        "spreadsheet",
                },
            }

        finally:
            workbook.close()

    def _format_value(
        self,
        value,
    ) -> str:

        if value is None:
            return ""

        if isinstance(
            value,
            bool,
        ):
            return (
                "TRUE"
                if value
                else "FALSE"
            )

        return str(
            value
        ).strip()